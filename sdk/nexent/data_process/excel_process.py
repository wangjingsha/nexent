import io
import os
from copy import deepcopy

import openpyxl


def process_excel_file(file_path: str, chunking_strategy: str = "basic", **params) -> list:
    """
    Process an Excel file and return its contents in a structured format
    
    Args:
        file_path: Path to the Excel file
        chunking_strategy: Strategy for chunking (not used in Excel processing)
        **params: Additional parameters
        
    Returns:
        List of processed chunks with metadata
    """
    processor = ExcelProcessor(input_file=file_path)
    raw_content = processor.process()
    
    # Convert raw content to proper chunk format
    chunks = []
    for i, content_text in enumerate(raw_content):
        chunk = {
            "content": content_text,
            "path_or_url": file_path,
            "filename": os.path.basename(file_path),
            "metadata": { # Keep original metadata nested if needed, but ensure top-level fields are present
                "chunk_index": i,
                "source_type": "excel",
                "file_type": "xlsx" if file_path.lower().endswith('.xlsx') else "xls"
            }
        }
        chunks.append(chunk)
    
    return chunks


class ExcelProcessor:
    def __init__(self, input_file=None, file_data=None):
        """
        Initialize Excel processor
        :param input_file: Excel file path (optional)
        :param file_data: Excel file data stream (optional, can be binary data or file object)
        Must provide either input_file or file_data parameter
        Check data source and load file
        Check if file exists
        Load from file path
        Load from data stream
        If binary data, use io.BytesIO to wrap
        If file object (e.g., object returned by open() or io.BytesIO)
        file_data must be binary data or file object
        Use openpyxl to load data stream
        Create deep copy
        Failed to load Excel file
        """
        if input_file is None and file_data is None:
            raise ValueError("Must provide input_file or file_data parameter")

        self.input_file = input_file
        self.file_data = file_data

        # Check data source and load file
        try:
            if self.input_file:
                # Check if file exists
                if not self._check_file_exists():
                    raise FileNotFoundError(f"File does not exist or is inaccessible: {self.input_file}")
                # Load from file path
                self.wb_original = openpyxl.load_workbook(self.input_file)
            else:
                # Load from data stream
                if isinstance(file_data, bytes):
                    # If binary data, use io.BytesIO to wrap
                    file_obj = io.BytesIO(file_data)
                elif hasattr(file_data, 'read'):
                    # If file object (e.g., object returned by open() or io.BytesIO)
                    file_obj = file_data
                else:
                    raise ValueError("file_data must be binary data or file object")

                # Use openpyxl to load data stream
                self.wb_original = openpyxl.load_workbook(file_obj)

            # Create deep copy
            self.wb_copy = deepcopy(self.wb_original)
        except Exception as e:
            raise Exception(f"Failed to load Excel file: {str(e)}")

    def process(self):
        """
        Process Excel file and return content

        Args:
            output_file (str, optional): Output Excel file path

        Returns:
            tuple: (list of processed content, output file path)
        """
        contents = []
        for sheet_name in self.wb_original.sheetnames:
            sheet = self.wb_original[sheet_name]
            sheet_copy = self.wb_copy[sheet_name]
            # If the sheet has only one column of content, merge the column content
            if self._is_single_column(sheet):
                content = self._get_content_one_col(sheet, sheet_name)
            else:
                begin_row, max_col = self._get_title_row(sheet_copy)
                self._merge_all_cells(sheet_copy)

                title_key = self._get_title_key(begin_row, sheet)
                remark = self._get_remark(sheet, begin_row)
                content = self._get_excel_content(title_key, remark, sheet, begin_row, sheet_name)
            contents.extend(content)
        return contents

    def _generate_output_filename(self):
        """Generate output filename"""
        file_dir, file_name = os.path.split(self.input_file)
        name, ext = os.path.splitext(file_name)
        return os.path.join(file_dir, f"{name}_unmerged{ext}")

    def _get_title_row(self, sheet):
        """Get header row information"""
        max_col = 0
        position_max_col = 0

        self._merge_columns(sheet)
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            non_empty_cells = sum(1 for cell in row if cell.value is not None)

            if non_empty_cells > max_col:
                max_col = non_empty_cells
                position_max_col = row_idx
        return position_max_col, max_col

    def _is_single_column(self, sheet):
        """Determine if there is only one column of content"""
        _, max_col = self._get_title_row(sheet)
        return max_col < 2

    def _get_content_one_col(self, sheet, sheet_name):
        """Process sheet with only one column of content"""
        content_str = ""
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if any(cell is not None for cell in row):
                content_str = content_str + row[0].replace("\n", "<br>") + "\n"
        return [content_str + "\n————" + sheet_name]

    def _get_remark(self, sheet, begin_row):
        """Get remarks before the header"""
        remark = ""
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None for cell in row):
                continue
            if row_idx >= begin_row:
                break
            remark = remark + "<br>" + self._join_tuple_elements(row)
        return remark

    def _get_excel_content(self, title_key, remark, sheet, begin_row, sheet_name):
        """Get table content"""

        def _get_content_row(title_key, row, remark):
            if remark:
                title_key.append("Remarks before the header")
                row = row + (remark,)

            result = {str(k).replace("\n", "<br>"): str(v).replace("\n", "<br>") for k, v in zip(title_key, row)}
            return self._dict_to_markdown_table(result) + "\n————" + sheet_name

        content = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None for cell in row):
                continue
            if row_idx <= begin_row:
                continue
            content.append(_get_content_row(title_key, row, remark))
        return content

    def _check_file_exists(self):
        """
        Check if file exists and is readable
        :return: bool
        """
        return os.path.isfile(self.input_file) and os.access(self.input_file, os.R_OK)

    @staticmethod
    def _dict_to_markdown_table(data):
        """Convert dictionary to Markdown table"""
        keys = []
        values = []
        for key, value in data.items():
            keys.append("None" if key is None else str(key))
            values.append("None" if key is None else str(value))
        table = "| " + " | ".join(keys) + " |\n"
        table += "| " + " | ".join(["---"] * len(keys)) + " |\n"
        table += "| " + " | ".join(map(str, values)) + " |"
        return table

    @staticmethod
    def _join_tuple_elements(input_tuple):
        """Join non-None elements in tuple with semicolon"""
        return ";".join(str(item) for item in input_tuple if item is not None)

    def _merge_columns(self, sheet):
        """Handle column merging"""
        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            if str(merged_range)[0] != str(merged_range)[3]:
                continue
            sheet.unmerge_cells(str(merged_range))

        for merged_range in merged_ranges:
            if str(merged_range)[0] != str(merged_range)[3]:
                continue

            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = sheet.cell(row=min_row, column=min_col).value

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value

    def _get_title_key(self, begin_row, sheet):
        """Get header keys"""
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None for cell in row):
                continue
            if row_idx == begin_row:
                return list(row)
        return []

    def _merge_all_cells(self, sheet):
        """Merge all cells"""
        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))

        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = sheet.cell(row=min_row, column=min_col).value

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value
